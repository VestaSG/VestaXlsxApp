#!/usr/bin/env python3

import codecs
import sys
sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer)

print("Content-type: text/html; charset=utf-8\n")

# https://github.com/huntflow/api/blob/master/ru/user.md
# /me не возвращает id пользователя

import requests
import json
import copy

class HFAPI:
	def __init__(self, token):
		self.set_token(token)
		self.host = 'https://api.huntflow.ru/'
		self.heads = {'Authorization': 'Bearer ' + self.token, 'User-Agent': 'Python3. VestaXlsxApp (test@huntflow.ru)'}
		self.tosave = dict()
		self.key_field = "id" # имя id в ответе API

	def set_token(self, token):
		self.token = token

	def set_uid(self, uid):
		self.uid = uid

	def load(self, id = 0):
		return len(self.j)

	def id_by_field(self, name, field):
		if len(self.j):
			for ent in self.j:
				if ent[field] == name:
					return ent[self.key_field]

	def request(self, way, json_string):
		r = requests.post(self.host + way, headers=self.heads, data=json_string)

	def save(self):
		return False


class Candidat(HFAPI):
	def set_vac(self, v):
		self.tosave["vacancy"] = v # тут INT (id вакансии)

	def set_name(self, v):
		newstrar = []
		newstr = ""
		for i in v: # TODO: обратить внимание на корректность работы цикла
			if i == ' ':
				newstrar.append(newstr)
				newstr = ""
			else:
				newstr = newstr + i
		newstrar.append(newstr)

		self.tosave["first_name"] = newstrar[1]
		self.tosave["last_name"] = newstrar[0]
		try:
			self.tosave["middle_name"] = newstrar[2]
		except:
			pass

	def set_step(self, v):
		self.tosave["status"] = v # INT (id статуса)

	def set_comment(self, v):
		self.tosave["comment"] = v

	def set_money(self, v):
		self.tosave["money"] = str(v)

	def set_fid(self, v):
		if v:
			self.tosave["files"] = []
			f = dict()
			f["id"] = v
			self.tosave["files"].append(f)

	def upd_vac_step_comment(self, id):
		mini = dict()
		mini["comment"] = self.tosave["comment"]
		mini["status"] = self.tosave["status"]
		mini["vacancy"] = self.tosave["vacancy"]
		try:
			mini["files"] = self.tosave["files"]
		except:
			pass
		r = requests.post(self.host + 'account/' + str(self.uid) + "/applicants/" + str(id) + "/vacancy", headers=self.heads, json=mini)

	def add_entity(self):
		# отбросить поля vac_step_comment
		mini = copy.deepcopy(self.tosave)
		del mini["comment"]
		del mini["status"]
		del mini["vacancy"]
		try:
			del mini["files"]
		except:
			pass
		r = requests.post(self.host + 'account/' + str(self.uid) + '/applicants', headers=self.heads, json=mini)
		j = json.loads(r.text)
		id = j[self.key_field]
		# TODO: загрузка файла резюме
		self.upd_vac_step_comment(id)
		return id


class Step(HFAPI):
	def load(self, id = 0):
		r = requests.get(self.host + 'account/' + str(self.uid) + '/vacancy/statuses', headers=self.heads)
		j = json.loads(r.text)
		self.j = j['items']
		return super().load()

	def id_by_name(self, name):
		return self.id_by_field(name, "name")


class Vacancy(HFAPI):
	def load(self, id = 0):
		r = requests.get(self.host + 'account/' + str(self.uid) + '/vacancies', headers=self.heads)
		j = json.loads(r.text)
		self.j = j['items']
		return super().load()

	def id_by_name(self, name):
		return self.id_by_field(name, "position")

class File(HFAPI):
	def set_vac_name(self, name):
		self.vac_nm = name

	def set_cand_name(self, name):
		self.cand_nm = name

	def new_file(self, cand_nm, vac_nm):
		# TODO: пути должны быть абсолютные
		self.set_cand_name(cand_nm)
		self.set_vac_name(vac_nm)
		name = "Тестовое задание/" + self.vac_nm + "/"+ self.cand_nm
		ext = ".doc"
		mim = "application/msword"
		i = 0
		while i < 2:
			if self.open_file(name + ext):
				fcontent = self.f.read()
				post = {'file': ('f' + ext, fcontent, mim)}
				r = requests.post(self.host + 'account/' + str(self.uid) + '/upload', headers=self.heads, files=post)
				j = json.loads(r.text)
				return j['id']
			ext = ".pdf"
			mim = "application/pdf"
			i += 1
		return False

	def open_file(self, name):
		try:
			self.f = open(name.encode('utf-8'), 'rb')
		except IOError:
			return False
		return True


from openpyxl import load_workbook # https://pypi.org/project/openpyxl/
# openpyxl uses https://pypi.org/project/jdcal/
class XLSLoader:
	def __init__(self, name):
		wb = load_workbook(name)
		self.sheet = wb.active
		lc = self.sheet.max_column
		for i in range(1, lc+1): # перечисление ячеек заголовка (row=1)
			col = self.sheet.cell(row=1, column=i).value
			if "Должность" == col:
				self.vac_col = i
			elif "ФИО" == col:
				self.fio_col = i
			elif "Ожидания по ЗП" == col:
				self.zp_col = i
			elif "Комментарий" == col:
				self.comment_col = i
			elif "Статус" == col:
				self.step_col = i

	def row_count(self):
		return self.sheet.max_row - 1

	def out_vac(self, rowid):
		return self.sheet.cell(row=rowid, column=self.vac_col).value

	def out_fio(self, rowid):
		return self.sheet.cell(row=rowid, column=self.fio_col).value

	def out_zp(self, rowid):
		return self.sheet.cell(row=rowid, column=self.zp_col).value

	def out_comment(self, rowid):
		return self.sheet.cell(row=rowid, column=self.comment_col).value

	def out_step(self, rowid):
		return self.sheet.cell(row=rowid, column=self.step_col).value

class VestaXlsxApp:
	def __init__(self, file):
		self.load_conf()
		self.xl = XLSLoader(file)
		self.step = Step(self.token)
		self.step.set_uid(self.uid)
		self.vac = Vacancy(self.token)
		self.vac.set_uid(self.uid)

	def load(self):
		self.step.load()
		self.vac.load()
		for i in range(2, 2+self.xl.row_count()): # range(2, row_count + 1)
			f_obj = File(self.token)
			f_obj.set_uid(self.uid)
			fid = f_obj.new_file(self.xl.out_fio(i), self.xl.out_vac(i))

			c_obj = Candidat(self.token)
			c_obj.set_uid(self.uid)

			c_obj.set_fid(fid)
			c_obj.set_name(self.xl.out_fio(i))
			c_obj.set_money(self.xl.out_zp(i))
			c_obj.set_comment(self.xl.out_comment(i))
			c_obj.set_step(self.step.id_by_name(self.xl.out_step(i)))
			c_obj.set_vac(self.vac.id_by_name(self.xl.out_vac(i)))
			c_obj.add_entity()
			del c_obj

	def load_conf(self):
		f = open('conf.json')
		j = json.loads(f.read())
		self.token = j["token"]
		self.uid = j["uid"]


# print(r.request.headers)

core_obj = VestaXlsxApp("testbase.xlsx")
core_obj.load()
print("Ok")
